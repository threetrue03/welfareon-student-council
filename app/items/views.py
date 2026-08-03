from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import OperationalError, transaction
from django.db.models.deletion import ProtectedError
from django.db.models import Count, Q
from django.http import FileResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST

from accounts.models import User
from dashboard.google_sheets import schedule_google_sheet_sync
from backup_utils import create_backup_archive, recent_backup_archives
from accounts.permissions import AdminCsvWriteError, admin_required, get_admin_ids, is_admin_student_id, set_admin_membership, write_admin_ids
from rentals.models import ConsumableIssueRecord, RentalRecord, ReturnRecord
from students.models import Student

from .forms import CategoryForm, ItemForm
from .models import Category, EquipmentUnit, Item

PAGE_SIZE = 10
PAGE_WINDOW_SIZE = 8


def _admin_csv_write_failed(request, error, redirect_to='items:worker_list'):
    messages.error(request, str(error))
    if transaction.get_connection().in_atomic_block:
        transaction.set_rollback(True)
    return redirect(redirect_to)


def _pagination_query_string(request):
    params = request.GET.copy()
    params.pop('page', None)
    return params.urlencode()


def _attach_page_window(page_obj, window_size=PAGE_WINDOW_SIZE):
    total_pages = page_obj.paginator.num_pages
    if total_pages <= window_size:
        start_page = 1
        end_page = total_pages
    else:
        start_page = page_obj.number - (window_size // 2) + 1
        end_page = start_page + window_size - 1
        if start_page < 1:
            start_page = 1
            end_page = window_size
        if end_page > total_pages:
            end_page = total_pages
            start_page = total_pages - window_size + 1
    page_obj.page_window = range(start_page, end_page + 1)
    return page_obj


def _paginate(request, queryset, page_size=PAGE_SIZE):
    paginator = Paginator(queryset, page_size)
    page_obj = paginator.get_page(request.GET.get('page'))
    return _attach_page_window(page_obj)


def _cell_to_text(value):
    if value is None:
        return ''
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_item_type(value):
    text = str(value or '').strip().lower()
    if text in {'비품', 'equipment', 'equip', 'e'}:
        return Item.ItemType.EQUIPMENT
    if text in {'소모품', 'consumable', 'consume', 'c'}:
        return Item.ItemType.CONSUMABLE
    return ''


def _normalize_admin_value(value):
    text = _cell_to_text(value).strip().lower()
    return text in {'관리자', 'admin', 'administrator', 'true', 't', 'yes', 'y', '1', 'o', 'on', '권한 있음'}


def _normalize_fee_status(value):
    text = _cell_to_text(value).strip().lower()
    if not text:
        return Student.FeeStatus.PAID
    if text in {'납부', '납부자', 'paid', 'pay', 'p', 'true', 't', 'yes', 'y', '1', 'o', 'on'}:
        return Student.FeeStatus.PAID
    if text in {'미납', '미납자', 'unpaid', 'u', 'false', 'f', 'no', 'n', '0', 'x', 'off'}:
        return Student.FeeStatus.UNPAID
    if text in {'환불', '환불자', 'refunded', 'refund', 'r'}:
        return Student.FeeStatus.REFUNDED
    return ''


def _backup_before_mutation(request, reason):
    try:
        backup_path = create_backup_archive(reason)
        messages.info(request, f'데이터 변경 전 자동 백업을 생성했습니다: {backup_path.name}')
        return backup_path
    except Exception as exc:
        messages.error(request, f'자동 백업 생성에 실패했습니다. 기존 데이터를 보호하기 위해 작업을 중단합니다: {exc}')
        return None


def _parse_quantity(value, *, row_number, field_name, errors, required=True):
    if value is None or str(value).strip() == '':
        if required:
            errors.append(f'{row_number}행: {field_name}을 입력해주세요.')
        return None
    try:
        if isinstance(value, float) and not value.is_integer():
            raise ValueError
        quantity = int(value)
    except (TypeError, ValueError):
        errors.append(f'{row_number}행: {field_name}은 정수로 입력해주세요.')
        return None
    if quantity < 0:
        errors.append(f'{row_number}행: {field_name}은 0 이상으로 입력해주세요.')
        return None
    return quantity


DB_LOCKED_MESSAGE = '현재 이전 작업을 정리하는 중입니다. 잠시 후 다시 업로드해주세요. 문제가 반복되면 프로그램을 완전히 종료한 뒤 다시 실행해주세요.'


def _is_database_locked_error(error):
    return 'database is locked' in str(error).lower()


def _build_import_preview(rows):
    return [{
        'name': row.get('name', ''),
        'location': row.get('location', ''),
        'category': row.get('category', ''),
        'item_type_label': '비품' if row.get('item_type') == Item.ItemType.EQUIPMENT else '소모품',
        'total_quantity': row.get('total_quantity', ''),
        'current_quantity': '' if row.get('current_quantity') is None else row.get('current_quantity'),
    } for row in rows[:8]]


def sync_equipment_units(item, new_total):
    existing_numbers = set(item.units.values_list('number', flat=True))
    for number in range(1, new_total + 1):
        if number not in existing_numbers:
            EquipmentUnit.objects.create(item=item, number=number)
    item.units.filter(number__gt=new_total).update(status=EquipmentUnit.Status.INACTIVE)


def get_filtered_items(request):
    categories = Category.objects.annotate(item_count=Count('items')).order_by('name')
    items = Item.objects.select_related('category').prefetch_related('units').order_by('item_type', 'category__name', 'name')
    query = request.GET.get('q', '').strip()
    item_type = request.GET.get('type', '').strip()
    category_id = request.GET.get('category', '').strip()
    if query:
        items = items.filter(Q(name__icontains=query) | Q(location__icontains=query) | Q(category__name__icontains=query))
    if item_type in dict(Item.ItemType.choices):
        items = items.filter(item_type=item_type)
    if category_id.isdigit():
        items = items.filter(category_id=category_id)
    return categories, items, query, item_type, category_id


def build_item_list_context(request, *, item_form=None, active_modal=''):
    categories, items, query, item_type, category_id = get_filtered_items(request)
    page_obj = _paginate(request, items)
    return {
        'active_tab': 'db',
        'db_sub_tab': 'items',
        'categories': categories,
        'items': page_obj.object_list,
        'page_obj': page_obj,
        'pagination_query_string': _pagination_query_string(request),
        'query': query,
        'selected_type': item_type,
        'selected_category': category_id,
        'item_type_choices': Item.ItemType.choices,
        'status_choices': EquipmentUnit.Status.choices,
        'item_form': item_form or ItemForm(initial={'item_type': Item.ItemType.EQUIPMENT}),
        'active_modal': active_modal,
    }


def build_category_list_context(request, *, category_form=None, active_modal=''):
    query = request.GET.get('q', '').strip()
    categories = Category.objects.annotate(item_count=Count('items')).order_by('name')
    if query:
        categories = categories.filter(name__icontains=query)
    page_obj = _paginate(request, categories)
    return {
        'active_tab': 'db',
        'db_sub_tab': 'categories',
        'categories': page_obj.object_list,
        'page_obj': page_obj,
        'pagination_query_string': _pagination_query_string(request),
        'query': query,
        'category_form': category_form or CategoryForm(),
        'active_modal': active_modal,
    }


def _worker_queryset(request):
    query = request.GET.get('q', '').strip()
    role_filter = request.GET.get('role', '').strip()
    workers = User.objects.order_by('-is_active', 'student_id')
    if query:
        workers = workers.filter(Q(name__icontains=query) | Q(student_id__icontains=query))
    admin_ids = get_admin_ids()
    if role_filter == 'admin':
        workers = workers.filter(student_id__in=admin_ids)
    elif role_filter == 'worker':
        workers = workers.exclude(student_id__in=admin_ids)
    return workers, query, role_filter, admin_ids


def build_worker_list_context(request):
    workers, query, role_filter, admin_ids = _worker_queryset(request)
    page_obj = _paginate(request, workers)
    worker_rows = []
    for worker in page_obj.object_list:
        worker.is_csv_admin = str(worker.student_id) in admin_ids
        worker_rows.append(worker)
    return {
        'active_tab': 'db',
        'db_sub_tab': 'workers',
        'workers': worker_rows,
        'page_obj': page_obj,
        'pagination_query_string': _pagination_query_string(request),
        'query': query,
        'selected_role': role_filter,
    }



def _is_worker_admin(worker):
    return is_admin_student_id(worker.student_id)


def _target_admin_password_is_valid(request, worker):
    password = request.POST.get('target_admin_password', '')
    return bool(password and worker.check_password(password))


def _require_target_admin_password(request, worker):
    if not _is_worker_admin(worker):
        return True
    if _target_admin_password_is_valid(request, worker):
        return True
    messages.error(request, '관리자 계정을 수정/삭제하려면 해당 관리자 비밀번호를 입력해야 합니다.')
    return False


@login_required
@admin_required
def item_list(request):
    return render(request, 'items/item_list.html', build_item_list_context(request))


@login_required
@admin_required
def category_list(request):
    return render(request, 'items/category_list.html', build_category_list_context(request))


@login_required
@admin_required
@transaction.atomic
def item_create(request):
    if request.method != 'POST':
        return redirect('items:item_list')
    form = ItemForm(request.POST)
    if form.is_valid():
        item = form.save(commit=False)
        if item.item_type == Item.ItemType.EQUIPMENT:
            item.current_quantity = None
        item.save()
        if item.item_type == Item.ItemType.EQUIPMENT:
            sync_equipment_units(item, item.total_quantity)
        messages.success(request, '물품이 추가되었습니다.')
        schedule_google_sheet_sync('물품 추가')
        return redirect('items:item_list')
    return render(request, 'items/item_list.html', build_item_list_context(request, item_form=form, active_modal='item'))


@login_required
@admin_required
@transaction.atomic
def item_update(request, pk):
    item = get_object_or_404(Item.objects.prefetch_related('units'), pk=pk)
    old_total = item.total_quantity
    old_type = item.item_type
    if request.method == 'POST':
        form = ItemForm(request.POST, instance=item)
        if form.is_valid():
            item = form.save(commit=False)
            if item.item_type == Item.ItemType.EQUIPMENT:
                item.current_quantity = None
            item.save()
            if item.item_type == Item.ItemType.EQUIPMENT:
                if old_type != Item.ItemType.EQUIPMENT:
                    item.units.all().delete()
                sync_equipment_units(item, item.total_quantity)
                for unit in item.units.all():
                    status = request.POST.get(f'unit_status_{unit.id}')
                    if status in dict(EquipmentUnit.Status.choices):
                        unit.status = status
                        unit.save(update_fields=['status', 'updated_at'])
            else:
                item.units.all().delete()
            messages.success(request, '물품 정보가 수정되었습니다.')
            schedule_google_sheet_sync('물품 수정')
            return redirect('items:item_list')
    return redirect('items:item_list')


@login_required
@admin_required
@require_POST
def item_delete(request, pk):
    item = get_object_or_404(Item, pk=pk)
    item_name = item.name
    linked_record_count = (
        RentalRecord.objects.filter(item=item).count()
        + ReturnRecord.objects.filter(item=item).count()
        + ConsumableIssueRecord.objects.filter(item=item).count()
    )
    if linked_record_count:
        messages.error(request, f'{item_name} 물품은 대여/반납/소모품 지급 기록과 연결되어 있어 삭제할 수 없습니다. 기록 보존을 위해 삭제 대신 상태 변경 또는 수량 조정을 사용해주세요.')
        return redirect('items:item_list')
    if item.units.filter(status=EquipmentUnit.Status.BORROWED).exists():
        messages.error(request, f'{item_name} 물품은 현재 대여 중인 번호가 있어 삭제할 수 없습니다.')
        return redirect('items:item_list')
    try:
        item.delete()
    except ProtectedError:
        messages.error(request, f'{item_name} 물품은 기존 기록과 연결되어 있어 삭제할 수 없습니다.')
        return redirect('items:item_list')
    messages.success(request, f'{item_name} 물품이 삭제되었습니다.')
    schedule_google_sheet_sync('물품 삭제')
    return redirect('items:item_list')


@login_required
@admin_required
def category_create(request):
    if request.method != 'POST':
        return redirect('items:category_list')
    form = CategoryForm(request.POST)
    if form.is_valid():
        form.save()
        messages.success(request, '카테고리가 추가되었습니다.')
        schedule_google_sheet_sync('카테고리 추가')
        return redirect('items:category_list')
    return render(request, 'items/category_list.html', build_category_list_context(request, category_form=form, active_modal='category'))


@login_required
@admin_required
def category_update(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, '카테고리 정보가 수정되었습니다.')
            schedule_google_sheet_sync('카테고리 수정')
            return redirect('items:category_list')
    return redirect('items:category_list')


@login_required
@admin_required
@require_POST
def category_delete(request, pk):
    category = get_object_or_404(Category.objects.annotate(item_count=Count('items')), pk=pk)
    category_name = category.name
    if category.item_count:
        messages.error(request, f'{category_name} 카테고리에 연결된 물품이 {category.item_count}개 있어 삭제할 수 없습니다. 먼저 물품의 카테고리를 변경해주세요.')
        return redirect('items:category_list')
    category.delete()
    messages.success(request, f'{category_name} 카테고리가 삭제되었습니다.')
    schedule_google_sheet_sync('카테고리 삭제')
    return redirect('items:category_list')


@login_required
@admin_required
def student_list(request):
    query = request.GET.get('q', '').strip()
    students = Student.objects.order_by('student_id')
    if query:
        students = students.filter(Q(name__icontains=query) | Q(student_id__icontains=query))
    page_obj = _paginate(request, students)
    return render(request, 'items/student_list.html', {
        'active_tab': 'db',
        'db_sub_tab': 'students',
        'students': page_obj.object_list,
        'page_obj': page_obj,
        'pagination_query_string': _pagination_query_string(request),
        'query': query,
        'fee_status_choices': Student.FeeStatus.choices,
    })


@login_required
@admin_required
@require_POST
def student_create(request):
    name = request.POST.get('name', '').strip()
    student_id = request.POST.get('student_id', '').strip()
    if not name or not student_id:
        messages.error(request, '학생 이름과 학번을 입력해주세요.')
    elif Student.objects.filter(student_id=student_id).exists():
        messages.error(request, '이미 등록된 학번입니다.')
    else:
        Student.objects.create(name=name, student_id=student_id, fee_status=Student.FeeStatus.PAID)
        messages.success(request, '학생이 추가되었습니다.')
        schedule_google_sheet_sync('학생 추가')
    return redirect('items:student_list')


@login_required
@admin_required
@require_POST
def student_update(request, pk):
    student = get_object_or_404(Student, pk=pk)
    name = request.POST.get('name', '').strip()
    student_id = request.POST.get('student_id', '').strip()
    fee_status = request.POST.get('fee_status', Student.FeeStatus.PAID)
    try:
        overdue_count = int(request.POST.get('overdue_count', 0))
    except (TypeError, ValueError):
        overdue_count = 0
    if overdue_count < 0:
        overdue_count = 0
    if not name or not student_id:
        messages.error(request, '학생 이름과 학번을 입력해주세요.')
    elif Student.objects.filter(student_id=student_id).exclude(pk=student.pk).exists():
        messages.error(request, '이미 등록된 학번입니다.')
    else:
        student.name = name
        student.student_id = student_id
        if fee_status in dict(Student.FeeStatus.choices):
            student.fee_status = fee_status
        student.overdue_count = overdue_count
        student.is_blacklisted = request.POST.get('is_blacklisted') == 'on'
        if not student.is_blacklisted:
            student.blacklist_until = None
        student.save()
        messages.success(request, '학생 정보가 수정되었습니다.')
        schedule_google_sheet_sync('학생 수정')
    return redirect('items:student_list')


@login_required
@admin_required
@require_POST
@transaction.atomic
def student_delete(request, pk):
    student = get_object_or_404(Student, pk=pk)
    active_records = RentalRecord.objects.select_related('item', 'unit').filter(student=student, status=RentalRecord.Status.ACTIVE)
    if active_records.exists():
        item_list = ', '.join(f'{record.item.name} {record.unit.number}번' for record in active_records)
        messages.error(request, f'현재 대여 중인 물품이 있어서 학생을 삭제할 수 없습니다. 먼저 반납 처리 후 삭제해주세요. 반납 필요 물품: {item_list}')
        return redirect('items:student_list')
    ConsumableIssueRecord.objects.filter(student=student).delete()
    ReturnRecord.objects.filter(student=student).delete()
    RentalRecord.objects.filter(student=student).delete()
    student_label = f'{student.student_id} {student.name}'
    student.delete()
    messages.success(request, f'{student_label} 학생이 삭제되었습니다.')
    schedule_google_sheet_sync('학생 삭제')
    return redirect('items:student_list')


@login_required
@admin_required
def worker_list(request):
    return render(request, 'items/worker_list.html', build_worker_list_context(request))


@login_required
@admin_required
@require_POST
@transaction.atomic
def worker_create(request):
    name = request.POST.get('name', '').strip()
    student_id = request.POST.get('student_id', '').strip()
    password = request.POST.get('password', '')
    is_admin = request.POST.get('is_admin') == 'on'
    if not name or not student_id or not password:
        messages.error(request, '이름, 학번, 비밀번호를 모두 입력해주세요.')
    elif User.objects.filter(student_id=student_id).exists():
        messages.error(request, '이미 등록된 근무자 학번입니다.')
    else:
        worker = User.objects.create_user(student_id=student_id, name=name, password=password)
        worker.visible_password = password
        worker.is_active = True
        worker.role = User.Role.ADMIN if is_admin else User.Role.WORKER
        worker.is_staff = is_admin
        worker.save(update_fields=['visible_password', 'is_active', 'role', 'is_staff'])
        try:
            set_admin_membership(student_id, is_admin)
        except AdminCsvWriteError as error:
            return _admin_csv_write_failed(request, error)
        messages.success(request, '근무자가 추가되었습니다.')
        schedule_google_sheet_sync('근무자 추가')
    return redirect('items:worker_list')


@login_required
@admin_required
@require_POST
@transaction.atomic
def worker_update(request, pk):
    worker = get_object_or_404(User, pk=pk)
    if not _require_target_admin_password(request, worker):
        return redirect('items:worker_list')

    old_student_id = worker.student_id
    name = request.POST.get('name', '').strip()
    student_id = request.POST.get('student_id', '').strip()
    password = request.POST.get('password', '')
    is_admin = request.POST.get('is_admin') == 'on'
    if not name or not student_id:
        messages.error(request, '이름과 학번을 모두 입력해주세요.')
    elif User.objects.filter(student_id=student_id).exclude(pk=worker.pk).exists():
        messages.error(request, '이미 등록된 근무자 학번입니다.')
    else:
        worker.name = name
        worker.student_id = student_id
        worker.role = User.Role.ADMIN if is_admin else User.Role.WORKER
        worker.is_staff = is_admin
        update_fields = ['name', 'student_id', 'role', 'is_staff']
        if password:
            worker.set_password(password)
            worker.visible_password = password
            update_fields.extend(['password', 'visible_password'])
        worker.save(update_fields=update_fields)
        try:
            set_admin_membership(student_id, is_admin, old_student_id=old_student_id)
        except AdminCsvWriteError as error:
            return _admin_csv_write_failed(request, error)
        messages.success(request, '근무자 정보가 수정되었습니다.')
        schedule_google_sheet_sync('근무자 수정')
    return redirect('items:worker_list')


@login_required
@admin_required
@require_POST
@transaction.atomic
def worker_delete(request, pk):
    worker = get_object_or_404(User, pk=pk)
    if worker.pk == request.user.pk:
        messages.error(request, '현재 로그인한 본인 계정은 삭제할 수 없습니다.')
        return redirect('items:worker_list')
    if not _require_target_admin_password(request, worker):
        return redirect('items:worker_list')

    worker_label = f'{worker.student_id} {worker.name}'
    admin_ids = get_admin_ids()
    if str(worker.student_id) in admin_ids and len(admin_ids) <= 1:
        messages.error(request, '관리자는 최소 1명 이상 남아 있어야 합니다. 마지막 관리자 계정은 비활성화할 수 없습니다.')
        return redirect('items:worker_list')
    try:
        set_admin_membership(worker.student_id, False)
    except AdminCsvWriteError as error:
        return _admin_csv_write_failed(request, error)
    worker.is_active = False
    worker.role = User.Role.WORKER
    worker.is_staff = False
    worker.save(update_fields=['is_active', 'role', 'is_staff'])
    messages.success(request, f'{worker_label} 근무자를 비활성화했습니다. 기존 대여/반납/근무 기록은 유지됩니다.')
    schedule_google_sheet_sync('근무자 비활성화')
    return redirect('items:worker_list')


@login_required
@admin_required
@require_POST
def worker_password_reveal(request, pk):
    worker = get_object_or_404(User, pk=pk)
    if _is_worker_admin(worker):
        if not _target_admin_password_is_valid(request, worker):
            return JsonResponse({'ok': False, 'message': '해당 관리자 비밀번호가 올바르지 않습니다.'}, status=403)
    return JsonResponse({
        'ok': True,
        'password': worker.visible_password or '',
        'message': '' if worker.visible_password else '저장된 표시용 비밀번호가 없습니다. 새 비밀번호로 변경하면 이후 확인할 수 있습니다.',
    })

@login_required
@admin_required
def data_backup(request):
    if request.method == 'POST':
        backup_path = create_backup_archive('manual')
        return FileResponse(open(backup_path, 'rb'), as_attachment=True, filename=backup_path.name)
    backups = [{'name': path.name, 'size': path.stat().st_size} for path in recent_backup_archives()]
    return render(request, 'items/data_backup.html', {'active_tab': 'db', 'db_sub_tab': 'data_backup', 'backups': backups})


@login_required
@admin_required
def db_import(request):
    errors = []
    parsed_rows = []
    preview_rows = []
    if request.method == 'POST':
        upload = request.FILES.get('xlsx_file')
        if not upload:
            errors.append('xlsx 파일을 선택해주세요.')
        elif not upload.name.lower().endswith('.xlsx'):
            errors.append('xlsx 파일만 업로드할 수 있습니다.')
        else:
            try:
                from openpyxl import load_workbook
                workbook = load_workbook(upload, data_only=True)
                sheet = workbook.active
            except Exception:
                errors.append('xlsx 파일을 읽을 수 없습니다. 파일 형식을 확인해주세요.')
            else:
                header_values = [_cell_to_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1), [])]
                has_location_column = len(header_values) >= 2 and '위치' in header_values[1]
                for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if has_location_column:
                        values = list(row[:6]) + [None] * (6 - len(row[:6]))
                        name = str(values[0] or '').strip()
                        location = str(values[1] or '').strip()
                        category_name = str(values[2] or '').strip()
                        item_type = _normalize_item_type(values[3])
                        total_raw = values[4]
                        current_raw = values[5]
                    else:
                        values = list(row[:5]) + [None] * (5 - len(row[:5]))
                        name = str(values[0] or '').strip()
                        location = ''
                        category_name = str(values[1] or '').strip()
                        item_type = _normalize_item_type(values[2])
                        total_raw = values[3]
                        current_raw = values[4]
                    if all(value is None or str(value).strip() == '' for value in values):
                        continue
                    total_quantity = _parse_quantity(total_raw, row_number=row_number, field_name='전체 수량', errors=errors)
                    current_quantity = _parse_quantity(current_raw, row_number=row_number, field_name='현재 수량', errors=errors, required=(item_type == Item.ItemType.CONSUMABLE))
                    if not name:
                        errors.append(f'{row_number}행: 물품명을 입력해주세요.')
                    if not category_name:
                        errors.append(f'{row_number}행: 카테고리를 입력해주세요.')
                    if not item_type:
                        errors.append(f'{row_number}행: 물품 특성은 비품 또는 소모품으로 입력해주세요.')
                    if item_type == Item.ItemType.EQUIPMENT and total_quantity is not None and total_quantity < 1:
                        errors.append(f'{row_number}행: 비품의 전체 수량은 1 이상이어야 합니다.')
                    if item_type == Item.ItemType.CONSUMABLE and total_quantity is not None and current_quantity is not None and current_quantity > total_quantity:
                        errors.append(f'{row_number}행: 소모품 현재 수량은 전체 수량보다 클 수 없습니다.')
                    parsed_rows.append({'row_number': row_number, 'name': name, 'location': location, 'category': category_name, 'item_type': item_type, 'total_quantity': total_quantity, 'current_quantity': current_quantity})
                if not parsed_rows:
                    errors.append('가져올 물품 데이터가 없습니다. 2행부터 데이터를 입력해주세요.')
                if not errors:
                    backup_path = _backup_before_mutation(request, 'inventory-import')
                    if backup_path is None:
                        preview_rows = _build_import_preview(parsed_rows)
                        return render(request, 'items/db_import.html', {'active_tab': 'db', 'db_sub_tab': 'inventory_import', 'errors': errors, 'preview_rows': preview_rows})
                    try:
                        with transaction.atomic():
                            ConsumableIssueRecord.objects.all().delete()
                            ReturnRecord.objects.all().delete()
                            RentalRecord.objects.all().delete()
                            EquipmentUnit.objects.all().delete()
                            Item.objects.all().delete()
                            Category.objects.all().delete()
                            category_cache = {}
                            for row in parsed_rows:
                                category = category_cache.get(row['category'])
                                if category is None:
                                    category, _ = Category.objects.get_or_create(name=row['category'])
                                    category_cache[row['category']] = category
                                item = Item.objects.create(name=row['name'], location=row.get('location', ''), category=category, item_type=row['item_type'], total_quantity=row['total_quantity'], current_quantity=None if row['item_type'] == Item.ItemType.EQUIPMENT else row['current_quantity'])
                                if item.item_type == Item.ItemType.EQUIPMENT:
                                    sync_equipment_units(item, item.total_quantity)
                    except OperationalError as error:
                        if _is_database_locked_error(error):
                            errors.append(DB_LOCKED_MESSAGE)
                        else:
                            raise
                    else:
                        messages.success(request, f'{len(parsed_rows)}개 물품을 xlsx 파일 기준으로 새로 가져왔습니다.')
                        schedule_google_sheet_sync('물품 DB 가져오기')
                        return redirect('items:db_import')
                preview_rows = _build_import_preview(parsed_rows)
    return render(request, 'items/db_import.html', {'active_tab': 'db', 'db_sub_tab': 'inventory_import', 'errors': errors, 'preview_rows': preview_rows})


@login_required
@admin_required
def student_db_import(request):
    errors = []
    preview_rows = []
    parsed_rows = []
    if request.method == 'POST':
        upload = request.FILES.get('xlsx_file')
        if not upload:
            errors.append('xlsx 파일을 선택해주세요.')
        elif not upload.name.lower().endswith('.xlsx'):
            errors.append('xlsx 파일만 업로드할 수 있습니다.')
        else:
            try:
                from openpyxl import load_workbook
                workbook = load_workbook(upload, data_only=True)
                sheet = workbook.active
            except Exception:
                errors.append('xlsx 파일을 읽을 수 없습니다. 파일 형식을 확인해주세요.')
            else:
                seen_student_ids = set()
                for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    values = list(row[:3]) + [None] * (3 - len(row[:3]))
                    if all(_cell_to_text(value) == '' for value in values):
                        continue
                    student_id = _cell_to_text(values[0])
                    name = _cell_to_text(values[1])
                    fee_status = _normalize_fee_status(values[2])
                    if not student_id:
                        errors.append(f'{row_number}행: 학번을 입력해주세요.')
                    if not name:
                        errors.append(f'{row_number}행: 이름을 입력해주세요.')
                    if not fee_status:
                        errors.append(f'{row_number}행: 학생회비 납부 여부는 납부, 미납, 환불 중 하나로 입력해주세요.')
                    if student_id:
                        if student_id in seen_student_ids:
                            errors.append(f'{row_number}행: 학번 {student_id}이 파일 안에서 중복되었습니다.')
                        seen_student_ids.add(student_id)
                    parsed_rows.append({'row_number': row_number, 'student_id': student_id, 'name': name, 'fee_status': fee_status, 'fee_status_label': dict(Student.FeeStatus.choices).get(fee_status, '')})
                if not parsed_rows:
                    errors.append('가져올 학생 데이터가 없습니다. 2행부터 데이터를 입력해주세요.')
                preview_rows = parsed_rows[:10]
                if not errors:
                    backup_path = _backup_before_mutation(request, 'student-import')
                    if backup_path is None:
                        return render(request, 'items/student_db_import.html', {'active_tab': 'db', 'db_sub_tab': 'student_import', 'errors': errors, 'preview_rows': preview_rows})
                    try:
                        with transaction.atomic():
                            ConsumableIssueRecord.objects.all().delete()
                            ReturnRecord.objects.all().delete()
                            RentalRecord.objects.all().delete()
                            EquipmentUnit.objects.filter(status=EquipmentUnit.Status.BORROWED).update(status=EquipmentUnit.Status.AVAILABLE)
                            Student.objects.all().delete()
                            for row in parsed_rows:
                                Student.objects.create(student_id=row['student_id'], name=row['name'], fee_status=row['fee_status'])
                    except OperationalError as error:
                        if _is_database_locked_error(error):
                            errors.append(DB_LOCKED_MESSAGE)
                        else:
                            raise
                    else:
                        messages.success(request, f'{len(parsed_rows)}명의 학생 명단을 새로 가져왔습니다.')
                        schedule_google_sheet_sync('학생 DB 가져오기')
                        return redirect('items:student_db_import')
    return render(request, 'items/student_db_import.html', {'active_tab': 'db', 'db_sub_tab': 'student_import', 'errors': errors, 'preview_rows': preview_rows})


@login_required
@admin_required
def worker_db_import(request):
    errors = []
    preview_rows = []
    parsed_rows = []
    if request.method == 'POST':
        upload = request.FILES.get('xlsx_file')
        if not upload:
            errors.append('xlsx 파일을 선택해주세요.')
        elif not upload.name.lower().endswith('.xlsx'):
            errors.append('xlsx 파일만 업로드할 수 있습니다.')
        else:
            try:
                from openpyxl import load_workbook
                workbook = load_workbook(upload, data_only=True)
                sheet = workbook.active
            except Exception:
                errors.append('xlsx 파일을 읽을 수 없습니다. 파일 형식을 확인해주세요.')
            else:
                seen_student_ids = set()
                existing_worker_ids = set(User.objects.values_list('student_id', flat=True))
                for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    values = list(row[:4]) + [None] * (4 - len(row[:4]))
                    if all(_cell_to_text(value) == '' for value in values):
                        continue
                    student_id = _cell_to_text(values[0])
                    name = _cell_to_text(values[1])
                    password = _cell_to_text(values[2])
                    is_admin = _normalize_admin_value(values[3])
                    if not student_id:
                        errors.append(f'{row_number}행: 학번을 입력해주세요.')
                    if not name:
                        errors.append(f'{row_number}행: 이름을 입력해주세요.')
                    if not password and student_id not in existing_worker_ids:
                        errors.append(f'{row_number}행: 신규 근무자는 비밀번호를 입력해주세요.')
                    if student_id:
                        if student_id in seen_student_ids:
                            errors.append(f'{row_number}행: 학번 {student_id}이 파일 안에서 중복되었습니다.')
                        seen_student_ids.add(student_id)
                    parsed_rows.append({'row_number': row_number, 'student_id': student_id, 'name': name, 'password': password, 'is_admin': is_admin})
                if not parsed_rows:
                    errors.append('가져올 근무자 데이터가 없습니다. 2행부터 데이터를 입력해주세요.')
                preview_rows = parsed_rows[:10]

                imported_ids = {row['student_id'] for row in parsed_rows if row.get('student_id')}
                admin_ids = {row['student_id'] for row in parsed_rows if row.get('student_id') and row.get('is_admin')}
                current_admin_id = str(request.user.student_id)
                if not admin_ids:
                    errors.append('관리자 계정이 최소 1명 이상 있어야 합니다. 관리자 여부가 1/관리자인 행을 포함해주세요.')
                if current_admin_id not in imported_ids:
                    errors.append(f'현재 로그인한 관리자({current_admin_id})가 근무자 DB 파일에 없습니다. 접속 불가를 막기 위해 현재 관리자 행을 포함해야 합니다.')
                elif current_admin_id not in admin_ids:
                    errors.append(f'현재 로그인한 관리자({current_admin_id})의 관리자 여부가 꺼져 있습니다. 현재 관리자는 관리자 권한을 유지해야 합니다.')

                if not errors:
                    backup_path = _backup_before_mutation(request, 'worker-import')
                    if backup_path is None:
                        return render(request, 'items/worker_db_import.html', {'active_tab': 'db', 'db_sub_tab': 'worker_import', 'errors': errors, 'preview_rows': preview_rows})
                    try:
                        with transaction.atomic():
                            User.objects.exclude(student_id__in=list(imported_ids)).update(is_active=False)
                            current_password_preserved = False
                            for row in parsed_rows:
                                worker, created = User.objects.get_or_create(student_id=row['student_id'], defaults={'name': row['name']})
                                worker.name = row['name']
                                worker.role = User.Role.ADMIN if row['is_admin'] else User.Role.WORKER
                                worker.is_staff = row['is_admin']
                                worker.is_active = True
                                update_fields = ['name', 'role', 'is_staff', 'is_active']
                                if row['password']:
                                    if worker.pk == request.user.pk:
                                        current_password_preserved = True
                                    else:
                                        worker.visible_password = row['password']
                                        worker.set_password(row['password'])
                                        update_fields.extend(['visible_password', 'password'])
                                worker.save(update_fields=update_fields)
                            write_admin_ids(admin_ids)
                            if current_password_preserved:
                                messages.info(request, '현재 로그인 중인 관리자 계정의 비밀번호는 세션 보호를 위해 기존 값으로 유지했습니다.')
                    except OperationalError as error:
                        if _is_database_locked_error(error):
                            errors.append(DB_LOCKED_MESSAGE)
                        else:
                            raise
                    except AdminCsvWriteError as error:
                        return _admin_csv_write_failed(request, error, redirect_to='items:worker_db_import')
                    else:
                        messages.success(request, f'{len(parsed_rows)}명의 근무자 DB를 새로 가져왔습니다. 관리자 {len(admin_ids)}명, 비활성화 {User.objects.filter(is_active=False).count()}명입니다.')
                        schedule_google_sheet_sync('근무자 DB 가져오기')
                        return redirect('items:worker_db_import')
    return render(request, 'items/worker_db_import.html', {'active_tab': 'db', 'db_sub_tab': 'worker_import', 'errors': errors, 'preview_rows': preview_rows})
